"""
Evaluate predicted exam answers stored as JSON files against an official answer
sheet (CSV or XLSX).

Usage examples
--------------
python evaluate_answers.py --answers 2025_kMLE_answers.csv --predictions-dir json
python evaluate_answers.py --answers 2025_kMLE_answers.xlsx --predictions-dir json
"""
from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Dict, Iterable, Optional, Tuple

ChoiceKey = Tuple[str, int]

_CIRCLED_TO_DIGIT = {
    "①": "1",
    "②": "2",
    "③": "3",
    "④": "4",
    "⑤": "5",
}


def _normalize_choice(raw: Optional[object]) -> Optional[str]:
    """Convert various answer notations to a string digit 1-5.

    Returns None for blank or unrecognized values.
    """
    if raw is None:
        return None
    text = str(raw).strip()
    if not text:
        return None
    if text in _CIRCLED_TO_DIGIT:
        return _CIRCLED_TO_DIGIT[text]
    if text.isdigit() and text in {"1", "2", "3", "4", "5"}:
        return text
    return None


def load_answer_key(path: Path) -> Dict[ChoiceKey, str]:
    """Load the official answer key from a CSV or XLSX file.

    The file is expected to include the columns: 교시, 과목, 문제번호, 최종답안.
    The returned dictionary maps (subject, question_number) to the correct choice
    as a string digit between "1" and "5".
    """
    if not path.exists():
        raise FileNotFoundError(f"Answer file not found: {path}")

    ext = path.suffix.lower()
    if ext == ".csv":
        return _load_answer_key_csv(path)
    if ext in {".xlsx", ".xlsm"}:
        return _load_answer_key_xlsx(path)
    raise ValueError(f"Unsupported answer file format: {path.suffix}")


def _load_answer_key_csv(path: Path) -> Dict[ChoiceKey, str]:
    answer_key: Dict[ChoiceKey, str] = {}
    with path.open(newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f)
        for row in reader:
            subject = (row.get("과목") or "").strip()
            question_number = row.get("문제번호")
            choice = _normalize_choice(row.get("최종답안"))
            if not subject or question_number is None or choice is None:
                continue
            try:
                qnum = int(str(question_number).strip())
            except ValueError:
                continue
            answer_key[(subject, qnum)] = choice
    return answer_key


def _load_answer_key_xlsx(path: Path) -> Dict[ChoiceKey, str]:
    try:
        from openpyxl import load_workbook  # type: ignore
    except ImportError as exc:  # pragma: no cover - dependency hint
        raise ImportError(
            "Reading XLSX files requires the 'openpyxl' package. Install it with "
            "`pip install openpyxl`."
        ) from exc

    wb = load_workbook(path, data_only=True)
    sheet = wb.active
    answer_key: Dict[ChoiceKey, str] = {}

    # Assume the first row is the header.
    headers = [str(cell.value).strip() if cell.value is not None else "" for cell in next(sheet.iter_rows(max_row=1))]
    header_index = {name: idx for idx, name in enumerate(headers)}
    required = ["과목", "문제번호", "최종답안"]
    if not all(key in header_index for key in required):
        raise ValueError("XLSX file is missing one of the required columns: 과목, 문제번호, 최종답안")

    for row in sheet.iter_rows(min_row=2):
        subject = row[header_index["과목"]].value
        question_number = row[header_index["문제번호"]].value
        choice_raw = row[header_index["최종답안"]].value

        choice = _normalize_choice(choice_raw)
        if subject is None or question_number is None or choice is None:
            continue
        try:
            qnum = int(str(question_number).strip())
        except ValueError:
            continue
        answer_key[(str(subject).strip(), qnum)] = choice
    return answer_key


def load_predictions(path: Path) -> Optional[Dict[ChoiceKey, str]]:
    """Load model predictions from a JSON file.

    The JSON is expected to contain a list or a dict with a "results" list. Each
    result entry should include a "subject" and an "answers" list, where each
    answer item includes a "question_number" and a selected option (e.g.,
    "chosen_index"). Unsupported shapes return None so the caller can skip
    unrelated JSON files.
    """
    data = json.loads(path.read_text(encoding="utf-8"))
    if isinstance(data, dict) and "results" in data:
        records = data["results"]
    elif isinstance(data, list):
        records = data
    else:
        return None

    predictions: Dict[ChoiceKey, str] = {}
    for record in records:
        if not isinstance(record, dict):
            continue
        subject = (record.get("subject") or "").strip()
        answers = record.get("answers")
        if not subject or not isinstance(answers, Iterable):
            continue
        for answer in answers:
            if not isinstance(answer, dict):
                continue
            qnum = answer.get("question_number")
            choice_raw = (
                answer.get("chosen_index")
                or answer.get("chosen_answer")
                or answer.get("answer")
                or answer.get("prediction")
            )
            choice = _normalize_choice(choice_raw)
            if qnum is None or choice is None:
                continue
            try:
                number = int(str(qnum).strip())
            except ValueError:
                continue
            predictions[(subject, number)] = choice
    return predictions if predictions else None


def evaluate(predictions: Dict[ChoiceKey, str], answer_key: Dict[ChoiceKey, str]):
    correct = 0
    for key, correct_choice in answer_key.items():
        if predictions.get(key) == correct_choice:
            correct += 1
    total = len(answer_key)
    incorrect = total - correct
    return correct, incorrect, total


def iter_prediction_files(directory: Path) -> Iterable[Path]:
    for path in sorted(directory.rglob("*.json")):
        if path.is_file():
            yield path


def main() -> None:
    parser = argparse.ArgumentParser(description="Evaluate predicted answers against an official answer key.")
    parser.add_argument("--answers", required=True, type=Path, help="Path to the answer CSV or XLSX file.")
    parser.add_argument("--predictions-dir", required=True, type=Path, help="Directory containing prediction JSON files.")
    args = parser.parse_args()

    answer_key = load_answer_key(args.answers)
    total_questions = len(answer_key)
    if total_questions == 0:
        raise SystemExit("No valid entries were found in the answer key.")

    print(f"Loaded answer key from {args.answers} ({total_questions} questions)\n")

    found_file = False
    for json_path in iter_prediction_files(args.predictions_dir):
        predictions = load_predictions(json_path)
        if predictions is None:
            continue
        found_file = True
        correct, incorrect, total = evaluate(predictions, answer_key)
        accuracy = correct / total * 100
        print(f"{json_path}:")
        print(f"  Correct:   {correct}")
        print(f"  Incorrect: {incorrect}")
        print(f"  Accuracy:  {accuracy:.2f}%\n")

    if not found_file:
        print("No prediction files with an 'answers' payload were found.")


if __name__ == "__main__":
    main()
